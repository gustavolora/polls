from .models import Respondent, SurveyRealized, Answer, AnswerOptions, Questions
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView
from django.http import HttpResponse
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.models import User
from .models import Commune, Surveys, District, Questions
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.db.models import Count, Avg
from openpyxl import Workbook
from django.utils import timezone

@never_cache
def loginview(request):
    logout(request)
    return render(request, 'login.html')
    


def isAdmin(user):
    return user.is_superuser

## handler errors 404, 500 and 403

def handler404notfound(request, exc):
    return render(request, '404.html' ,{'exception':exc }, status=404)


class Handler500view(TemplateView):
    template_name = "500.html"

    @classmethod
    def as_error_view(cls):

        v = cls.as_view()
        def view(request):
            r = v(request)
            r.render()
            return r
        return view

@never_cache
@login_required
def index(request):
    return render(request, 'index.html',
                  {'user': request.user})


@never_cache
@user_passes_test(isAdmin, login_url='index')
@login_required
def adminPage(request):
    return render(request, 'admin.html')


@never_cache
@user_passes_test(isAdmin, login_url='index')
@login_required
def adminsurveydetail(request, survey_id):
    survey_realized = get_object_or_404(Surveys, id=survey_id)
    questions = Questions.objects.filter(survey=survey_realized)
    data = []
    for question in questions:
        response_counts_by_option = []
        total_count = 0

        answer_options = AnswerOptions.objects.filter(question=question)
        for option in answer_options:
            count = Answer.objects.filter(
                questions=question, answeroptions=option).count()
            response_counts_by_option.append(
                {'option': option, 'count': count})
            total_count += count

        data.append({'question': question, 'response_counts_by_option':
                    response_counts_by_option, 'total_count': total_count})
    total_survey_realized = SurveyRealized.objects.filter(survey=survey_realized).count()
    for item in data:
        for option in item['response_counts_by_option']:
            if item['total_count'] > 0:
                option['percentage'] = option['count'] / \
                    item['total_count'] * 100
            else:
                option['percentage'] = 0

    return render(request, 'survey_result.html', {'data': data,'total_count':total_survey_realized, 'survey_realized': survey_realized})


@user_passes_test(isAdmin, login_url='index')
@login_required
def adminsurveys(request):
    surveys = Surveys.objects.all()
    survey_data = []
    for survey in surveys:
        numero_preguntas = Questions.objects.filter(survey=survey).count()
        survey_data.append(
            {'survey': survey, 'numero_preguntas': numero_preguntas})

    if surveys is None:
        return render(request, 'admin_surveys.html', {'error': 'No te han asignado encuestas'})
    else:
        return render(request, 'admin_surveys.html', {'surveys_data': survey_data})


@never_cache
@login_required
def surveys(request):
    surveys = Surveys.objects.all()
    survey_data = []
    for survey in surveys:
        numero_preguntas = Questions.objects.filter(survey=survey).count()
        survey_data.append(
            {'survey': survey, 'numero_preguntas': numero_preguntas})

    if surveys is None:
        return render(request, 'surveys.html', {'error': 'No te han asignado encuestas'})
    else:
        return render(request, 'surveys.html', {'surveys_data': survey_data})


@never_cache
@login_required
def video(request):
    return render(request, 'video.html')


@never_cache
@login_required
def surveydetail(request, survey_id):
    communes = Commune.objects.filter()
    survey = get_object_or_404(Surveys, id=survey_id)
    questions = Questions.objects.filter(
        survey=survey).prefetch_related('answeroptions_set')

    context = {
        'survey_id': survey_id,
        'communes': communes,
        'questions': questions
    }
    return render(request, 'survey_detail.html', context)


@csrf_exempt
@login_required
def setVideo(request):
    return render(request, 'video.html')


@login_required
def getDistrict(request):
    comuna_id = request.GET.get('comuna_id')
    distritos = District.objects.filter(
        commune_id=comuna_id).values('id', 'name')

    return JsonResponse(list(distritos), safe=False)


@csrf_protect
@login_required
def saveAnswers(request):
    if request.method == 'POST':
        nombre = request.POST.get('name')
        telefono = request.POST.get('phone')
        direccion = request.POST.get('direccion')
        recomendation = request.POST.get('recomentation')
        commune_id = request.POST.get('comunaSelect')
        district_id = request.POST.get('barrioSelect')
        survey_id = request.POST.get('surveys_id')
        duration = request.POST.get('duration')
        latitud = request.POST.get('latitude')
        longitud = request.POST.get('longitude')
        print(request.POST)
        survey = Surveys.objects.get(id=survey_id)
        respondent = Respondent(
            name=nombre, address=direccion, phone=telefono, recomendations=recomendation)
        respondent.save()
        if  latitud == '' and longitud == '':
            latitud = 0
            longitud = 0
        survey_realized = SurveyRealized(user=request.user,
                                         survey=survey,
                                         respondent=respondent,
                                         commune_id=commune_id,
                                         district_id=district_id,
                                         duration=duration,
                                         latitude=latitud,
                                         longitude=longitud

                                         )
        survey_realized.save()

        for key, value in request.POST.items():
            if key.startswith('pregunta'):
                pregunta_id = key.replace('pregunta', '')
                respuesta_id = value
                try:
                    pregunta = Questions.objects.get(id=int(pregunta_id))
                    print('pregunta:', pregunta_id)
                    respuesta = AnswerOptions.objects.get(id=int(respuesta_id))
                    print('respuesta encontrada exitosamente')
                    answer = Answer(surveyrealized=survey_realized,
                                    answeroptions=respuesta, questions=pregunta)
                    print('respuesta guardada exitosamente')
                    answer.save()
                except (Questions.DoesNotExist, AnswerOptions.DoesNotExist):
                    print(
                        'Error: Pregunta o respuesta no encontrada en la base de datos')
        context = {
            'survey_id': survey_id
        }
        return render(request, 'video.html', context)

    return redirect('index')


@never_cache
def signin(request):
    if request.method == 'GET':
        logout(request)
        print('entro aqui')
        return render(request, 'login.html')
    else:
        logout(request)
        username = request.POST['username']
        converted_username = username.lower()
        user = authenticate(
            request, username=converted_username, password=request.POST['password'])
        if user is None:
            messages.error(request, 'Usuario o contraseña incorrectos')
            return render(request, 'login.html')
        else:
            if user.is_superuser:
                login(request, user)
                return redirect('adminpage')
            else:
                login(request, user)
                return redirect('index')


@never_cache
@staff_member_required
@login_required
def listpollsters(request):
    users_data = User.objects.exclude(username='admin').annotate(survey_count=Count('surveyrealized'), average_duration=Avg(
        'surveyrealized__duration')/60).order_by('-survey_count').values('username', 'survey_count', 'average_duration')
    context = {
        'users_data': users_data
    }
    return render(request, 'pollster_list.html', context)


@never_cache
@staff_member_required
@login_required
def excelreport(request):
    workbook = Workbook()
    sheet = workbook.active
    now = timezone.now()
    format = now.strftime('%Y%m%d%H%M')

    headers = [
        'Fecha de Encuesta',
        'Comuna',
        'Barrio',
        'Nombre del Encuestado',
        'Teléfono',
        'Dirección',
        'Encuestador',
        'Latitud',
        'Longitud',
        'Recomendaciones',
        'Duracion de la encuesta (Minutos)'
    ]
    
    

    survey_realizeds = SurveyRealized.objects.select_related(
        'respondent', 'commune', 'district', 'user'
    ).prefetch_related('answer_set')
    row = 2
    questions = Questions.objects.filter(
        survey__surveyrealized__isnull=False).distinct()
    for question in questions:
        headers.append(question.question)

    for column, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column).value = header


    for survey_realized in survey_realizeds:
        sheet.cell(row=row, column=1).value = survey_realized.date.strftime(
            '%Y-%m-%d')
        sheet.cell(row=row, column=2).value = survey_realized.commune.name
        sheet.cell(row=row, column=3).value = survey_realized.district.name
        sheet.cell(row=row, column=4).value = survey_realized.respondent.name
        sheet.cell(row=row, column=5).value = survey_realized.respondent.phone
        sheet.cell(row=row, column=6).value = survey_realized.respondent.address
        sheet.cell(row=row, column=7).value = survey_realized.user.username
        sheet.cell(row=row, column=8).value = survey_realized.latitude
        sheet.cell(row=row, column=9).value = survey_realized.longitude
        sheet.cell(row=row, column=10).value = survey_realized.respondent.recomendations
        duration_minutes = survey_realized.duration / 60
        sheet.cell(row=row, column=11).value = duration_minutes

        respondent_answers = Answer.objects.filter(
            surveyrealized__respondent=survey_realized.respondent)
        for respondent_answer in respondent_answers:
            for column, question in enumerate(questions, start=11):
                if respondent_answer.questions == question:
                    sheet.cell(
                        row=row, column=column).value = respondent_answer.answeroptions.options

        row += 1
    filename = format
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    workbook.save(response)
    return response


@csrf_exempt
@never_cache
@login_required
def signout(request):
    logout(request)
    return redirect('login')
